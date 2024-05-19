import pandas as pd
from datetime import datetime, timedelta
import numpy_financial as npf
from loan_function import calculate_playdate
from loan_function import *
from openpyxl.styles import Font, PatternFill

file_path = 'Loan IRR Calc Model.xlsx'
sheet_name1 = 'Prepay'
sheet_name2 = "Charged Off"

df_prepay = pd.read_excel(file_path,sheet_name=sheet_name1, header = 1) # second row as header
df_prepay = df_prepay.map(lambda x: f'{x:.17f}%' if isinstance(x, float) else x)

df_charge = pd.read_excel(file_path,sheet_name=sheet_name2)
df_charge = df_charge.map(lambda x: f'{x:.17f}%' if isinstance(x, float) else x)

pd.options.display.float_format = '{:.2f}'.format

df = pd.DataFrame({'Months': range(1, 38)})
df['Payment_Count'] = df['Months'].apply(lambda x: 0 if x == 1 else x - 1)
df['Playdate'] = df.apply(calculate_playdate, axis=1)
df['Scheduled_Principal'] = df['Months'].apply(calculate_scheduled_principal)
df['Scheduled_Interest'] = df.apply(lambda row: calculate_scheduled_interest(row['Months'], row['Scheduled_Principal']), axis=1)
df['Scheduled_Balance'] = df['Months'].apply(calculate_scheduled_balance, args=(df,))
df['Prepay_Speed'] = df.apply(calculate_prepay, axis=1, args=(df_prepay,))
df['Default_Rate'] = df.apply(calculate_default_rate, axis=1, args=(df_charge,))

# Initialize necessary columns with zeros
df['Recovery'] = 0.0
df['Servicing_CF'] = 0.0
df['Earnout_CF'] = 0.0
df['Balance'] = 0.0
df['Principal'] = 0.0
df['Default'] = 0.0
df['Prepay'] = 0.0
df['Interest_Amount'] = 0.0
df['Total_CF'] = 0.0

# Calculate values for each month
for i in range(len(df)):
    if df.loc[i, 'Months'] == 1:
        df.loc[i, 'Prepay'] = 0
        df.loc[i, 'Default'] = 0
        df.loc[i, 'Recovery'] = 0
        df.loc[i, 'Principal'] = 0
        df.loc[i, 'Balance'] = invested
        df.loc[i, 'Servicing_CF'] = 0
        df.loc[i, 'Earnout_CF'] = 0
        df.loc[i, 'Interest_Amount'] = 0
        df.loc[i, 'Total_CF'] = -invested * (1 + purchase_premium)
    else:
        df.loc[i, 'Prepay'] = (df.loc[i-1, 'Balance'] - ((df.loc[i-1, 'Balance'] - df.loc[i, 'Scheduled_Interest']) / df.loc[i-1, 'Scheduled_Balance'] * df.loc[i, 'Scheduled_Principal'])) * (df.loc[i, 'Prepay_Speed'])/100 * prepay_multiplier
        df.loc[i, 'Default'] = (df.loc[i-1, 'Balance'] * df.loc[i-1, 'Default_Rate'])/100 * default_multiplier
        df.loc[i, 'Recovery'] = df.loc[i, 'Default'] * recovery_rate
        df.loc[i, 'Principal'] = ((df.loc[i-1, 'Balance'] - df.loc[i, 'Default']) / df.loc[i-1, 'Scheduled_Balance'] * df.loc[i, 'Scheduled_Principal']) + df.loc[i, 'Prepay']
        df.loc[i, 'Balance'] = df.loc[i-1, 'Balance'] - df.loc[i, 'Default'] - df.loc[i, 'Principal']
        df.loc[i, 'Servicing_CF'] = (df.loc[i-1, 'Balance'] - df.loc[i-1, 'Default']) * (service_fee/12)
        df.loc[i, 'Earnout_CF'] = 0 if not (df.loc[i, 'Months'] in [13, 19]) else (earnout_fee / 2) * invested
        df.loc[i, 'Interest_Amount'] = (df.loc[i-1, 'Balance'] - df.loc[i, 'Default']) * (coupon_rate / 12)
        df.loc[i, 'Total_CF'] = (df.loc[i, 'Principal'] + df.loc[i, 'Interest_Amount'] + df.loc[i, 'Recovery'] - df.loc[i, 'Servicing_CF'] - df.loc[i, 'Earnout_CF'])

cash_flows = df['Total_CF'].to_list()

irr = npf.irr(cash_flows)*12

# print(f"The Internal Rate of Return (IRR) is: {irr:.2%}")

df = df.round(2)
df = df.map(lambda x: 0 if isinstance(x, float) and abs(x) < 1e-10 else x)

irr_df = pd.DataFrame({'IRR': [f'{irr:.2%}']})
with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=0, startcol=0)

    # Add the IRR to the next column
    irr_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=0, startcol=len(df.columns) + 1)

    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Set the width of the columns
    for column in worksheet.columns:
        max_length = 0
        column_name = column[0].column_letter  # Get the column name
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_name].width = adjusted_width

    # Set the font style for the header
    header_font = Font(bold=True)
    for cell in worksheet[1]:  # Assuming the first row is the header
        cell.font = header_font

    # Bold and highlight the IRR header
    irr_header_cell = worksheet.cell(row=1, column=len(df.columns) + 2)  # Adjusted to match the position
    irr_header_cell.font = Font(bold=True)

    # Bold and highlight the IRR value
    irr_value_cell = worksheet.cell(row=2, column=len(df.columns) + 2)  # Adjusted to match the position
    irr_value_cell.font = Font(bold=True)
    irr_value_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Highlight in yellow

print("Exported df to output.xlsx with formatting.")